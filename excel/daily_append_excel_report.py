from typing import Dict

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def main(full_path , all_results: Dict[str, Dict], header_names, index, day_header):

    wb = load_workbook(full_path)

    STATUS_THRESHOLD = 70.0
    ws = wb.active
    ws.title = "Daily Check Report"

    # --- 스타일 정의 ---
    bold_font = Font(bold=True, name='Calibri', size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws['B2'] = f"CPU Over usage({STATUS_THRESHOLD}%) is fail / Memory : Memory Over usage ({STATUS_THRESHOLD}%) is fail"
    ws['B2'].font = Font(bold=True, size=12)


    # B3~B5 셀 병합 및 'Date' 텍스트 설정
    ws.merge_cells('B3:B5')
    date_header_cell = ws['B3']
    date_header_cell.value = "Date"
    date_header_cell.font = bold_font
    date_header_cell.alignment = center_align

    # --- 동적 헤더 및 데이터 생성 ---
    start_col = 3
    data_row = 6+index



    for display_name, internal_name in header_names.items():
        block_end_col = start_col + 9
        cpu_start_col, cpu_end_col = start_col, start_col + 4
        mem_start_col, mem_end_col = start_col + 5, start_col + 9

        ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=block_end_col)
        cell = ws.cell(row=3, column=start_col, value=display_name)
        cell.font = Font(bold=True, size=12)
        cell.alignment = center_align

        ws.merge_cells(start_row=4, start_column=cpu_start_col, end_row=4, end_column=cpu_end_col)
        cell = ws.cell(row=4, column=cpu_start_col, value="CPU Usage")
        cell.font = bold_font
        cell.alignment = center_align

        ws.merge_cells(start_row=4, start_column=mem_start_col, end_row=4, end_column=mem_end_col)
        cell = ws.cell(row=4, column=mem_start_col, value="Memory Usage")
        cell.font = bold_font
        cell.alignment = center_align


        detailed_headers = ["Min\n(%)", "Max\n(%)", "Avg\n(%)", "Status\n(Pass/Fail)", "Special Note"]

        for i, header in enumerate(detailed_headers * 2):
            cell = ws.cell(row=5, column=start_col + i, value=header)
            cell.font = bold_font
            cell.alignment = center_align

        cpu_data = all_results.get("CPU Usage", {}).get(internal_name, {})

        if cpu_data:
            ws.cell(row=data_row, column=cpu_start_col, value=round(cpu_data.get('min', 0), 3))
            ws.cell(row=data_row, column=cpu_start_col + 1, value=round(cpu_data.get('max', 0), 3))
            ws.cell(row=data_row, column=cpu_start_col + 2, value=round(cpu_data.get('avg', 0), 3))
            status_cell = ws.cell(row=data_row, column=cpu_start_col + 3)

            if cpu_data.get('max', 0) > STATUS_THRESHOLD:
                status_cell.value, status_cell.fill = "Fail", fail_fill
            else:
                status_cell.value, status_cell.fill = "Pass", pass_fill

            ws.cell(row=data_row, column=cpu_start_col + 4, value="이상없음")

        mem_data = all_results.get("Memory Usage", {}).get(internal_name, {})

        if mem_data:
            ws.cell(row=data_row, column=mem_start_col, value=round(mem_data.get('min', 0), 3))
            ws.cell(row=data_row, column=mem_start_col + 1, value=round(mem_data.get('max', 0), 3))
            ws.cell(row=data_row, column=mem_start_col + 2, value=round(mem_data.get('avg', 0), 3))
            status_cell = ws.cell(row=data_row, column=mem_start_col + 3)

            if mem_data.get('max', 0) > STATUS_THRESHOLD:
                status_cell.value, status_cell.fill = "Fail", fail_fill
            else:
                status_cell.value, status_cell.fill = "Pass", pass_fill

            ws.cell(row=data_row, column=mem_start_col + 4, value="이상없음")

        start_col = block_end_col + 1

    # B6 행에 실제 날짜
    date_value_cell = ws.cell(row=data_row, column=2, value=day_header[0:10])
    date_value_cell.font = bold_font
    date_value_cell.alignment = center_align

    # --- 최종 스타일 ---
    for col in range(2, start_col):
        ws.column_dimensions[get_column_letter(col)].width = 15
        for row in range(3, data_row + 1):
            ws.cell(row, col).border = thin_border
    ws.column_dimensions['B'].width = 12


    wb.save(full_path)

    print(f"\n✅ 엑셀 보고서가 성공적으로 생성되었습니다: {full_path}")